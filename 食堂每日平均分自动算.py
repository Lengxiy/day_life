import pandas as pd
import openpyxl
import os



data = pd.read_excel("./test1.xlsx", sheet_name="投票结果", na_values=['',0])
wb = openpyxl.load_workbook("./test1.xlsx")
ws = wb.get_sheet_by_name("投票结果")

#分值替换，都是5  10  15  20
data["饭菜新鲜度？"] = data["饭菜新鲜度？"].replace({"满意":20,"比较满意":15,"一般":10,"差":5})
data["饭菜品种如何？"] = data["饭菜品种如何？"].replace({"满意":20,"比较满意":15,"一般":10,"差":5})
data["饭菜味道如何？"] = data["饭菜味道如何？"].replace({"满意":20,"比较满意":15,"一般":10,"差":5})
data["食堂工作人员服务态度？"] = data["食堂工作人员服务态度？"].replace({"满意":20,"比较满意":15,"一般":10,"差":5})
data["食堂工作人员个人卫生（有否未带口罩、留长指甲等）？"] = data["食堂工作人员个人卫生（有否未带口罩、留长指甲等）？"].replace({"满意":20,"比较满意":15,"一般":10,"差":5})

# 只选择F列至K列的数据
df_subset = data.loc[:, "饭菜新鲜度？":"食堂工作人员个人卫生（有否未带口罩、留长指甲等）？"]

# 对每一行求和，并添加到新的列中
data["总分"] = df_subset.apply(lambda x: x.sum(), axis=1)
#对总分求和

total = data["总分"].sum()
#求平均分
average = total/(len(data["总分"]) - 1)
# 获取数据的行数和列数，并赋值给rows和cols变量
rows, cols = data.shape

data.to_excel("./test1.xlsx", sheet_name="投票结果")

# 选择要填充的单元格，指定row=1, column=cols+1
cell = ws.cell(row=1, column=cols+1)
cell1 = ws.cell(row=2, column=cols+1)
# 将average变量赋值给该单元格
cell.value = "平均分"
cell1.value = average

# 保存修改后的excle文件
wb.save("./test1.xlsx")